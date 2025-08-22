import base64
import json
import logging
import os
import re
import warnings

import openpyxl
from bs4 import XMLParsedAsHTMLWarning
from simplegmail import Gmail
from simplegmail.query import construct_query

# Ẩn cảnh báo BeautifulSoup XMLParsedAsHTMLWarning
warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

# Cấu hình logging -> ghi log ra file
logging.basicConfig(
    filename="mail_fetcher.log",
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)

gmail = Gmail()
os.makedirs("outputs", exist_ok=True)

downloaded_ids = set()


def load_allowed_subjects(excel_path="allowed_subjects.xlsx", sheet_name=None):
    """
    Đọc danh sách subject từ file Excel.
    Sẽ tìm cột có tên 'subject'.
    """
    if not os.path.exists(excel_path):
        logging.error(f"Không tìm thấy file {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    header = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    ]
    if "subject" not in header:
        logging.error("Không tìm thấy cột 'subject' trong file Excel.")
        return []

    subject_idx = header.index("subject")

    subjects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[subject_idx]:
            subjects.append(str(row[subject_idx]).strip())

    logging.info(f"Đã load {len(subjects)} ALLOWED_SUBJECTS từ {excel_path}")
    return subjects


def _save_file(target_dir, filename, file_data):
    """Hàm phụ để lưu file, tránh ghi đè nếu trùng tên."""
    os.makedirs(target_dir, exist_ok=True)

    base, ext = os.path.splitext(filename)
    counter = 1
    new_filename = filename

    while os.path.exists(os.path.join(target_dir, new_filename)):
        new_filename = f"{base}({counter}){ext}"
        counter += 1

    file_path = os.path.join(target_dir, new_filename)

    with open(file_path, "wb") as f:
        f.write(file_data)

    print(f"  Saved attachment: {file_path}")
    logging.info(f"Saved attachment: {file_path}")


def download_attachments_parent(service, user_id, msg_id, target_dir):
    """Lấy tất cả tệp đính kèm của 1 email."""
    try:
        message = service.users().messages().get(userId=user_id, id=msg_id).execute()
        payload = message.get("payload", {})

        def _extract_parts(parts):
            for part in parts:
                if part.get("filename"):
                    body = part.get("body", {})
                    att_id = body.get("attachmentId")
                    if att_id:
                        att = (
                            service.users()
                            .messages()
                            .attachments()
                            .get(userId=user_id, messageId=msg_id, id=att_id)
                            .execute()
                        )
                        data = att.get("data")
                        if data:
                            file_data = base64.urlsafe_b64decode(data.encode("UTF-8"))
                            _save_file(target_dir, part["filename"], file_data)

                if "parts" in part:
                    _extract_parts(part["parts"])

        if "parts" in payload:
            _extract_parts(payload["parts"])
    except Exception as e:
        logging.error(f"Lỗi tải file đính kèm từ mail {msg_id}: {e}")


def fetch_mails_in_date(allowed_subjects):
    query_params = {
        # "after": "2025/08/08", # có thể thêm giá trị này để xác định lấy mail từ khoảng thời gian nào. Nếu ko thì sẽ lấy tất cả các mail -> time sẽ lâu.
        "before": "2025/08/22",
    }
    query = construct_query(query_params)

    try:
        messages = gmail.get_messages(query=query)
        logging.info(f"Tìm thấy {len(messages)} email.")
    except Exception as e:
        logging.error(f"Lỗi khi fetch mail: {e}")
        return

    threads = {}
    for msg in messages:
        if msg.id in downloaded_ids:
            continue

        # Bỏ qua nếu subject ko thuộc 
        if msg.subject not in allowed_subjects:
            logging.info(f"Bỏ qua mail không hợp lệ: {msg.subject}")
            continue

        if msg.thread_id not in threads:
            threads[msg.thread_id] = []
        threads[msg.thread_id].append(msg)
        downloaded_ids.add(msg.id)

    for thread_id, msgs in threads.items():
        thread_data = []
        for msg in sorted(msgs, key=lambda x: x.date):
            mail_data = {
                "id": msg.id,
                "thread_id": thread_id,
                "from": msg.sender,
                "to": msg.recipient,
                "subject": msg.subject,
                "date": str(msg.date),
                "plain_text": msg.plain,
                "cc": msg.cc,
                "bcc": msg.bcc,
                "labels_ids": [label.name for label in msg.label_ids],
                "attachments": [att.filename for att in msg.attachments],
            }
            thread_data.append(mail_data)

        # Dữ liệu gồm thread_id + danh sách mail
        data_to_save = {"thread_id": thread_id, "messages": thread_data}

        safe_subject = re.sub(r'[\\/*?:"<>|]', "_", msgs[0].subject or "no_subject")
        thread_dir = os.path.join("outputs", f"{thread_id}_{safe_subject}")
        os.makedirs(thread_dir, exist_ok=True)

        # Lưu JSON đặt tên theo thread_id
        json_filename = f"{thread_id}.json"
        filename = os.path.join(thread_dir, json_filename)
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data_to_save, f, ensure_ascii=False, indent=4)

        print(f"Đã lưu hội thoại: {filename}")
        logging.info(f"Đã lưu hội thoại: {filename}")

        # Tải attachment vào cùng folder
        for msg in msgs:
            if msg.attachments:
                download_attachments_parent(gmail.service, "me", msg.id, thread_dir)


if __name__ == "__main__":
    subjects = load_allowed_subjects("allowed_subjects.xlsx") # Thay tên file excel chứa các giá trị subject
    if not subjects:
        print("Không có ALLOWED_SUBJECTS nào, thoát.")
    else:
        fetch_mails_in_date(subjects)
