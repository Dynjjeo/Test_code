import base64
import json
import logging
import os
import re
import warnings

import openpyxl
from bs4 import XMLParsedAsHTMLWarning
from google.auth.transport.requests import Request  # noqa: E402
from google.oauth2.credentials import Credentials  # noqa: E402
from google_auth_oauthlib.flow import InstalledAppFlow  # noqa: E402
from googleapiclient.discovery import build  # noqa: E402
from simplegmail.query import construct_query

# Import docling để đọc tệp đính kèm
try:
    from docling.datamodel.base_models import InputFormat
    from docling.document_converter import DocumentConverter

    DOCLING_AVAILABLE = True
    logging.info("Docling đã được import thành công")
except ImportError:
    DOCLING_AVAILABLE = False
    logging.warning("Docling không có sẵn. Cần cài đặt: pip install docling")

# Ẩn cảnh báo BeautifulSoup XMLParsedAsHTMLWarning
warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)

# Cấu hình logging -> ghi log ra file
logging.basicConfig(
    filename="mail_fetcher.log",
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)


downloaded_ids = set()

# Các định dạng file được hỗ trợ bởi docling
SUPPORTED_FORMATS = {
    ".pdf",
    ".docx",
    ".pptx",
    ".xlsx",
    ".html",
    ".md",
    ".txt",
    ".doc",
    ".ppt",
    ".xls",
    ".rtf",
    ".odt",
    ".odp",
    ".ods",
}


CREDENTIALS_FILE = os.getenv("WORKFLOWS_CLIENT_SECRET")
TOKEN_FILE = os.getenv("WORKFLOWS_GMAIL_TOKEN")


def init_gmail_service():
    SCOPES = [
        "https://www.googleapis.com/auth/gmail.settings.basic",
        "https://www.googleapis.com/auth/gmail.modify",
    ]

    creds = None

    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)

    # nếu chưa có gmail_token thì thực hiện
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=8080)

        with open(TOKEN_FILE, "w") as token:
            token.write(creds.to_json())

    # build gmail service
    service = build("gmail", "v1", credentials=creds)
    return service


def normalize_subject(subject: str) -> str:
    """
    Chuẩn hóa subject: loại bỏ prefix như 'Re:', 'Fwd:' và khoảng trắng thừa.
    """
    if not subject:
        return ""
    return re.sub(r"^(Re|Fwd):\s*", "", subject, flags=re.IGNORECASE).strip()


def load_allowed_subjects(excel_path, sheet_name=None, col_name=""):
    """
    Đọc danh sách subject từ file Excel.
    Sẽ tìm cột có tên 'subject'.
    """
    if not os.path.exists(excel_path):
        logging.error(f"Không tìm thấy file {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    header = [
        str(cell.value).strip().lower() if cell.value else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    ]
    if col_name not in header:
        logging.error("Không tìm thấy cột 'subject' trong file Excel.")
        return []

    subject_idx = header.index(col_name)

    subjects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[subject_idx]:
            subjects.append(normalize_subject(str(row[subject_idx]).strip()))

    logging.info(f"Đã load {len(subjects)} ALLOWED_SUBJECTS từ {excel_path}")
    return subjects


def _save_file(target_dir, filename, file_data):
    """Hàm phụ để lưu file, tránh ghi đè nếu trùng tên."""
    os.makedirs(target_dir, exist_ok=True)

    base, ext = os.path.splitext(filename)
    counter = 1
    new_filename = filename

    while os.path.exists(os.path.join(target_dir, new_filename)):
        new_filename = f"{base}({counter}){ext}"
        counter += 1

    file_path = os.path.join(target_dir, new_filename)

    with open(file_path, "wb") as f:
        f.write(file_data)

    print(f"Saved attachment: {file_path}")
    logging.info(f"Saved attachment: {file_path}")

    return file_path


def extract_content_with_docling(file_path):
    """
    Sử dụng docling để trích xuất nội dung từ tệp đính kèm
    """
    if not DOCLING_AVAILABLE:
        logging.warning("Docling không có sẵn, không thể trích xuất nội dung")
        return None

    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext not in SUPPORTED_FORMATS:
            logging.info(f"File {file_path} có định dạng không được hỗ trợ: {file_ext}")
            return None

        converter = DocumentConverter()
        result = converter.convert(file_path)
        content = result.document.export_to_markdown()

        logging.info(f"Đã trích xuất nội dung từ {file_path}")
        return content

    except Exception as e:
        logging.error(f"Lỗi khi trích xuất nội dung từ {file_path}: {e}")
        return None


def download_attachments_parent(service, user_id, msg_id, target_dir):
    """Lấy tất cả tệp đính kèm của 1 email và trích xuất nội dung."""
    attachments_info = []

    try:
        message = service.users().messages().get(userId=user_id, id=msg_id).execute()
        payload = message.get("payload", {})

        def _extract_parts(parts):
            for part in parts:
                if part.get("filename"):
                    body = part.get("body", {})
                    att_id = body.get("attachmentId")
                    if att_id:
                        att = (
                            service.users()
                            .messages()
                            .attachments()
                            .get(userId=user_id, messageId=msg_id, id=att_id)
                            .execute()
                        )
                        data = att.get("data")
                        if data:
                            file_data = base64.urlsafe_b64decode(data.encode("UTF-8"))
                            file_path = _save_file(
                                target_dir, part["filename"], file_data
                            )

                            attachment_context = extract_content_with_docling(file_path)

                            attachment_info = {
                                "filename": part["filename"],
                                "file_path": file_path,
                                "attachment_context": attachment_context
                                if attachment_context
                                else None,
                            }
                            attachments_info.append(attachment_info)

                if "parts" in part:
                    _extract_parts(part["parts"])

        if "parts" in payload:
            _extract_parts(payload["parts"])

    except Exception as e:
        logging.error(f"Lỗi tải file đính kèm từ mail {msg_id}: {e}")

    return attachments_info


def fetch_mails_in_date(allowed_subjects, after_default, before_default):
    query_params = {
        "after": after_default,
        "before": before_default,
    }
    query = construct_query(query_params)
    gmail = init_gmail_service()
    try:
        messages = gmail.get_messages(query=query)
        logging.info(f"Tìm thấy {len(messages)} email.")
    except Exception as e:
        logging.error(f"Lỗi khi fetch mail: {e}")
        return

    # Gom thread_id -> danh sách mail
    threads = {}
    valid_threads = set()

    for msg in messages:
        if msg.id in downloaded_ids:
            continue

        subj_norm = normalize_subject(msg.subject)

        # Nếu subject khớp danh sách allowed -> thread này được chấp nhận
        if subj_norm in allowed_subjects:
            valid_threads.add(msg.thread_id)

        threads.setdefault(msg.thread_id, []).append(msg)
        downloaded_ids.add(msg.id)

    # Chỉ xử lý những thread hợp lệ
    for thread_id, msgs in threads.items():
        if thread_id not in valid_threads:
            continue

        thread_data = []
        msgs_sorted = sorted(msgs, key=lambda x: x.date)

        for msg in msgs_sorted:
            attachments_info = []
            if msg.attachments:
                safe_subject = re.sub(
                    r'[\\/*?:"<>|]', "_", msgs_sorted[0].subject or "no_subject"
                )
                thread_dir = os.path.join("outputs", f"{thread_id}_{safe_subject}")
                attachments_info = download_attachments_parent(
                    gmail.service, "me", msg.id, thread_dir
                )

            mail_data = {
                "id": msg.id,
                "from": msg.sender,
                "to": msg.recipient,
                "subject": msg.subject,
                "date": str(msg.date),
                "plain_text": msg.plain,
                "cc": msg.cc,
                "bcc": msg.bcc,
                "labels_ids": [label.name for label in msg.label_ids],
                "attachments": attachments_info,
            }
            thread_data.append(mail_data)

        safe_subject = re.sub(
            r'[\\/*?:"<>|]', "_", msgs_sorted[0].subject or "no_subject"
        )
        thread_dir = os.path.join("outputs", f"{thread_id}_{safe_subject}")
        os.makedirs(thread_dir, exist_ok=True)

        data_to_save = {"thread_id": thread_id, "messages": thread_data}

        filename = os.path.join(thread_dir, f"{thread_id}.json")
        with open(filename, "w", encoding="utf-8") as f:
            json.dump(data_to_save, f, ensure_ascii=False, indent=4)

        print(f"Đã lưu hội thoại: {filename}")
        logging.info(f"Đã lưu hội thoại: {filename}")

        # Tạo file tổng hợp nội dung đính kèm
        all_attachments = []
        for msg_data in thread_data:
            if msg_data.get("attachments"):
                all_attachments.extend(msg_data["attachments"])

        if all_attachments:
            summary_filename = os.path.join(
                thread_dir, f"{thread_id}_attachment_summary.md"
            )
            with open(summary_filename, "w", encoding="utf-8") as f:
                f.write(f"# Tổng hợp nội dung đính kèm - Thread {thread_id}\n\n")
                f.write(f"**Chủ đề:** {msgs_sorted[0].subject}\n\n")

                for i, att_info in enumerate(all_attachments, 1):
                    f.write(f"## {i}. {att_info['filename']}\n\n")
                    f.write(f"**Đường dẫn file:** {att_info['file_path']}\n\n")
                    if att_info.get("attachment_context"):
                        f.write("**Nội dung:**\n\n")
                        f.write(att_info["attachment_context"])
                    else:
                        f.write(
                            "**Nội dung:** Không thể trích xuất nội dung từ file này.\n\n"
                        )
                    f.write("\n\n---\n\n")

            print(f"Đã tạo file tổng hợp: {summary_filename}")
            logging.info(f"Đã tạo file tổng hợp: {summary_filename}")


if __name__ == "__main__":
    subjects = load_allowed_subjects("allowed_subjects.xlsx")
    print(subjects)
    if not subjects:
        print("Không có ALLOWED_SUBJECTS nào, thoát.")
    else:
        fetch_mails_in_date(subjects)
